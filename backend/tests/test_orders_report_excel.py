from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.schemas.reports import (
    DailyReportDocument,
    SkuPerformance,
    StoreDailySummary,
    TrendPoint,
)
from app.services.reports.excel import write_daily_report_excel


def _document() -> DailyReportDocument:
    return DailyReportDocument(
        report_date=date(2026, 5, 31),
        store_summaries=[
            StoreDailySummary(
                seller_account_id=1,
                seller_name="hrm",
                marketplace_id="ATVPDKIKX0DER",
                currency="USD",
                ordered_product_sales=Decimal("50.00"),
                units_ordered=5,
                data_status="final",
            )
        ],
        totals={"ordered_product_sales": Decimal("50.00"), "units_ordered": Decimal("5")},
        warnings=[],
        data_source="orders",
        trend=[
            TrendPoint(
                period_label="M:2026-05",
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                currency="USD",
                ordered_product_sales=Decimal("50.00"),
                units_ordered=5,
                order_count=5,
            )
        ],
        sku_performance=[
            SkuPerformance(
                sku="SKU-1",
                asin="B0SKU1",
                product_name="Product SKU-1",
                currency="USD",
                units_ordered=5,
                ordered_product_sales=Decimal("50.00"),
            )
        ],
    )


def test_excel_has_trend_and_sku_sheets(tmp_path: Path) -> None:
    output = tmp_path / "report.xlsx"
    write_daily_report_excel(_document(), output)

    workbook = load_workbook(output)
    assert "Sales Trend" in workbook.sheetnames
    assert "SKU Performance" in workbook.sheetnames
